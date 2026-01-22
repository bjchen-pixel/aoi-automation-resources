#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
校正報告 Excel 生成器
使用前請安裝: pip install openpyxl
"""

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

def create_calibration_report():
    # 創建工作簿
    wb = Workbook()
    ws = wb.active
    ws.title = "校正報告"

    # 設置列寬
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 20

    # 邊框樣式
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # 標題
    ws['A1'] = '校正報告'
    ws['A1'].font = Font(size=18, bold=True)
    ws['A1'].alignment = Alignment(horizontal='center')
    ws.merge_cells('A1:F1')
    ws.row_dimensions[1].height = 30

    # 基本資料表
    row = 3
    ws[f'A{row}'] = '客戶:照敏企業股份有限公司'
    ws[f'C{row}'] = '校正者:慶揚儀器有限公司'
    ws[f'E{row}'] = '報告編號:107-016'

    row += 1
    ws[f'A{row}'] = '校正項目:長度'
    ws[f'C{row}'] = '校正儀器:金相顯微鏡'
    ws.merge_cells(f'C{row}:D{row}')

    row += 1
    ws[f'A{row}'] = '廠牌/型號:OLYMPUS/BH2'
    ws[f'C{row}'] = '溫度:25.0±1.0℃'
    ws[f'E{row}'] = '溼度:60.0±5.0% RH'

    row += 1
    ws[f'A{row}'] = '校正日期:2018/02/02'
    ws[f'C{row}'] = '有效期限:2019/02/01'
    ws[f'C{row}'].font = Font(color="FF0000")
    ws.merge_cells(f'C{row}:D{row}')

    # 為基本資料表添加邊框
    for r in range(3, 7):
        for c in range(1, 7):
            cell = ws.cell(row=r, column=c)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='left', vertical='center')

    # 標準器及附配儀器
    row = 8
    ws[f'A{row}'] = '標準器及附配儀器(詳 SGS 校正報告書)'
    ws[f'A{row}'].font = Font(bold=True)
    ws.merge_cells(f'A{row}:F{row}')

    row += 1
    headers = ['名稱', '廠牌/型號', '識別號碼', '校正日期', '下次校正日期']
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col)
        cell.value = header
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center', vertical='center')

    row += 1
    data = ['光學檢驗尺', 'Olympus/OB-MM', 'TAC2508617', '2017.08.17', '2018.08.16']
    for col, value in enumerate(data, start=1):
        cell = ws.cell(row=row, column=col)
        cell.value = value
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # 量測數據
    row = 12
    ws[f'A{row}'] = '量測數據'
    ws[f'A{row}'].font = Font(bold=True)
    ws.merge_cells(f'A{row}:F{row}')

    row += 1
    headers = ['參數/檔', '倍率', '標準值', '讀值', '誤差值', '備註']
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col)
        cell.value = header
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # 量測數據內容
    measurement_data = [
        ['標準刻度尺', '50X', '39.370mil', '39.318mil', '0.052mil', '上限 39.567 mil\n下限 39.173 mil'],
        ['標準刻度尺', '100X', '19.685mil', '19.659mil', '0.026mil', '上限 19.783 mil\n下限 19.587 mil'],
        ['標準刻度尺', '200X', '9.843mil', '9.829mil', '0.014mil', '上限 9.892 mil\n下限 9.794 mil'],
        ['標準刻度尺', '500X', '3.937mil', '3.931mil', '0.006mil', '上限 3.957 mil\n下限 3.917 mil'],
    ]

    for data_row in measurement_data:
        row += 1
        for col, value in enumerate(data_row, start=1):
            cell = ws.cell(row=row, column=col)
            cell.value = value
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        ws.row_dimensions[row].height = 30

    # 附註
    row += 2
    ws[f'A{row}'] = '※此量測校正的容許誤差為±十分之五，以上的量測值均在容許範圍內※'
    ws[f'A{row}'].font = Font(bold=True)
    ws.merge_cells(f'A{row}:F{row}')

    row += 2
    ws[f'A{row}'] = '附註：'
    ws[f'A{row}'].font = Font(bold=True)

    row += 1
    ws[f'A{row}'] = '上限 = 標準值 × 1.005'
    ws.merge_cells(f'A{row}:C{row}')

    row += 1
    ws[f'A{row}'] = '下限 = 標準值 × 0.995'
    ws.merge_cells(f'A{row}:C{row}')

    # 校驗者簽名
    row += 3
    ws[f'D{row}'] = '校驗者：'
    ws[f'E{row}'] = '_______________'

    # 保存文件
    filename = '校正報告_107-016.xlsx'
    wb.save(filename)
    print(f'✅ Excel 文件已生成：{filename}')

if __name__ == '__main__':
    try:
        create_calibration_report()
    except ImportError:
        print('❌ 請先安裝 openpyxl：')
        print('   pip install openpyxl')
    except Exception as e:
        print(f'❌ 生成失敗：{e}')
